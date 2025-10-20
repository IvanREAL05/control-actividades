import os
from datetime import datetime, date
from typing import Optional, Dict, Any
import logging
import json
from fastapi import APIRouter, HTTPException, Response, Query, Depends
from pydantic import BaseModel
import aiomysql
from cryptography.fernet import Fernet
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pytz
from io import BytesIO
from routes.ws_manager import manager
from routes.ws_manager_tabla import tabla_manager
# Importar la configuración de base de datos
from config.db import get_pool, fetch_one, fetch_all, execute_query, get_db_connection, get_pool

# Importar funciones de fecha
from utils.fecha import obtener_fecha_hora_cdmx, convertir_fecha_a_cdmx

# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Router FastAPI
router = APIRouter()
# Configuración de Fernet
FERNET_KEY = os.getenv('FERNET_KEY')
if not FERNET_KEY:
    raise ValueError("FERNET_KEY no está configurada en las variables de entorno")

fernet_cipher = Fernet(FERNET_KEY.encode())

# Constantes
FECHA_INICIO_CICLO = '2025-08-04'

# Modelos Pydantic
class EscaneoQRRequest(BaseModel):
    qr: str
    id_clase: int
    estado: str

class ActualizarAsistenciaRequest(BaseModel):
    matricula: str
    id_clase: int
    estado: str

class ActualizarEstadoRequest(BaseModel):
    id_estudiante: int
    id_clase: int
    estado: str

# Funciones utilitarias adaptadas
def obtener_fecha_hora_dict() -> Dict[str, str]:
    """Adapta la función de fecha para retornar un dict compatible con el código original"""
    fecha, hora, dia = obtener_fecha_hora_cdmx()
    return {
        'fecha': fecha.strftime('%Y-%m-%d'),
        'hora': hora.strftime('%H:%M:%S'),
        'dia': dia
    }

def obtener_dia_semana_texto(fecha_str: str) -> str:
    """Obtiene el día de la semana en texto español a partir de una fecha string"""
    fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
    # 🔧 CORREGIDO: Mismo array que obtener_fecha_hora_cdmx()
    dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    return dias[fecha_obj.weekday()]  # weekday() ya da 0=Lunes, 6=Domingo

# Función para inicializar asistencias
async def inicializar_asistencias(id_clase: int, connection: aiomysql.Connection):
    """Inicializa asistencias como 'ausente' para todos los estudiantes del grupo"""
    try:
        fecha_hoy = obtener_fecha_hora_cdmx()['fecha']
        
        async with connection.cursor() as cursor:
            # Obtener grupo de la clase
            await cursor.execute(
                "SELECT id_grupo FROM clase WHERE id_clase = %s",
                (id_clase,)
            )
            grupo_result = await cursor.fetchone()
            
            if not grupo_result:
                logger.warning("❌ Clase no encontrada para inicializar asistencias.")
                return
            
            id_grupo = grupo_result[0]
            
            # Obtener estudiantes del grupo
            await cursor.execute(
                "SELECT id_estudiante FROM estudiante WHERE id_grupo = %s",
                (id_grupo,)
            )
            estudiantes = await cursor.fetchall()
            
            for (id_estudiante,) in estudiantes:
                # Verificar si ya existe asistencia
                await cursor.execute(
                    "SELECT id_asistencia FROM asistencia WHERE id_estudiante = %s AND id_clase = %s AND fecha = %s",
                    (id_estudiante, id_clase, fecha_hoy)
                )
                check = await cursor.fetchone()
                
                if not check:
                    # Insertar como ausente
                    await cursor.execute(
                        "INSERT INTO asistencia (id_estudiante, id_clase, estado, fecha) VALUES (%s, %s, 'ausente', %s)",
                        (id_estudiante, id_clase, fecha_hoy)
                    )
                    logger.info(f"✅ Asistencia 'ausente' insertada para estudiante {id_estudiante}")
            
            await connection.commit()
            logger.info("🎉 Inicialización de asistencias completada")
            
    except Exception as error:
        logger.error(f"❌ Error en inicializar_asistencias: {error}")
        await connection.rollback()

# ENDPOINTS

@router.post("/")
async def escanear_qr(request: EscaneoQRRequest, connection: aiomysql.Connection = Depends(get_db_connection)):
    """Escanear código QR para registrar asistencia"""
    start_time = datetime.now()
    logger.info(f"📥 Petición recibida: {request.estado}, {request.id_clase}, QR recibido")

    try:
        # Desencriptar QR
        try:
            decrypted = fernet_cipher.decrypt(request.qr.encode()).decode()
        except Exception:
            raise HTTPException(status_code=400, detail="QR inválido o expirado")

        partes = [p.strip() for p in decrypted.split('|')]

        if len(partes) < 4:
            raise HTTPException(status_code=400, detail="Formato QR inválido")

        nombre_completo = partes[0]
        matricula = partes[1]
        grupo_texto = partes[2]

        # Obtener fecha y hora actual
        datos = obtener_fecha_hora_cdmx()
        fecha = datos["fecha"]
        hora = datos["hora"]
        hoy = fecha.strftime('%Y-%m-%d')
        hora_actual = hora.strftime('%H:%M:%S')

        async with connection.cursor() as cursor:
            # Consulta optimizada con JOINs
            consulta_completa = """
                SELECT 
                    e.id_estudiante,
                    g.id_grupo,
                    c.id_clase,
                    e.nombre,
                    e.apellido,
                    a.id_asistencia,
                    a.estado as estado_actual
                FROM estudiante e
                JOIN grupo g ON g.nombre = %s
                JOIN clase c ON c.id_grupo = g.id_grupo
                LEFT JOIN asistencia a ON a.id_estudiante = e.id_estudiante 
                    AND a.id_clase = c.id_clase 
                    AND a.fecha = %s
                WHERE e.matricula = %s
                    AND c.id_clase = %s 
                LIMIT 1
            """

            await cursor.execute(consulta_completa, (grupo_texto, hoy, matricula, request.id_clase))
            resultado = await cursor.fetchone()

            if not resultado:
                raise HTTPException(
                    status_code=404,
                    detail="Estudiante no encontrado o no hay clase activa para este grupo"
                )

            (id_estudiante, id_grupo, id_clase_datos, nombre, apellido,
             id_asistencia, estado_actual) = resultado

            # ✅ NUEVO: Obtener información de grupo y materia para WebSocket
            await cursor.execute("""
                SELECT g.nombre, m.nombre
                FROM grupo g
                JOIN clase c ON c.id_grupo = g.id_grupo
                JOIN materia m ON c.id_materia = m.id_materia
                WHERE c.id_clase = %s
            """, (request.id_clase,))
            info_clase = await cursor.fetchone()
            nombre_grupo, nombre_materia = info_clase if info_clase else ("", "")

            # ============================
            # Caso: ya existe asistencia
            # ============================
            if id_asistencia:
                if estado_actual == request.estado:
                    response_time = (datetime.now() - start_time).total_seconds() * 1000
                    logger.warning(f"⚠️ Escaneo repetido ({response_time:.0f}ms). Estado: {estado_actual}")
                    return {
                        "success": True,
                        "mensaje": f"{nombre} {apellido} ya estaba registrado como '{estado_actual}'",
                        "duplicado": True
                    }

                # Actualizar estado existente
                await cursor.execute(
                    "UPDATE asistencia SET estado = %s, hora_entrada = %s WHERE id_asistencia = %s",
                    (request.estado, hora_actual, id_asistencia)
                )
                await connection.commit()

                # 🔔 Difusión WebSocket - ✅ CON INFORMACIÓN COMPLETA
                mensaje_ws = json.dumps({
                    "evento": "asistencia_actualizada",
                    "id_estudiante": id_estudiante,
                    "nombre": nombre,
                    "apellido": apellido,
                    "id_clase": id_clase_datos,
                    "estado": request.estado,
                    "hora": hora_actual,
                    "fecha": hoy,
                    "nombre_grupo": nombre_grupo,
                    "nombre_materia": nombre_materia
                })
                
                logger.info(f"📡 Enviando WebSocket: {mensaje_ws[:150]}...")
                logger.info(f"👥 Clientes conectados: {len(manager.active_connections)}")
                await manager.broadcast(mensaje_ws)
                await tabla_manager.broadcast(mensaje_ws, id_clase=request.id_clase)

                response_time = (datetime.now() - start_time).total_seconds() * 1000
                logger.info(f"🔄 Asistencia actualizada ({response_time:.0f}ms): {nombre} {apellido} -> '{request.estado}'")
                return {
                    "success": True,
                    "mensaje": f"{nombre} {apellido} actualizado a '{request.estado}'",
                    "actualizado": True
                }

            # ============================
            # Caso: nueva asistencia
            # ============================
            await cursor.execute(
                "INSERT INTO asistencia (id_estudiante, id_clase, hora_entrada, estado, fecha) VALUES (%s, %s, %s, %s, %s)",
                (id_estudiante, request.id_clase, hora_actual, request.estado, hoy)
            )
            await connection.commit()

            # 🔔 Difusión WebSocket - ✅ CON INFORMACIÓN COMPLETA
            mensaje_ws = json.dumps({
                "evento": "nueva_asistencia",
                "id_estudiante": id_estudiante,
                "nombre": nombre,
                "apellido": apellido,
                "id_clase": request.id_clase,
                "estado": request.estado,
                "hora": hora_actual,
                "fecha": hoy,
                "nombre_grupo": nombre_grupo,
                "nombre_materia": nombre_materia
            })
            
            logger.info(f"📡 Enviando WebSocket: {mensaje_ws[:150]}...")
            logger.info(f"👥 Clientes conectados: {len(manager.active_connections)}")
            await manager.broadcast(mensaje_ws)
            await tabla_manager.broadcast(mensaje_ws, id_clase=request.id_clase)

            response_time = (datetime.now() - start_time).total_seconds() * 1000
            logger.info(f"✅ Asistencia registrada ({response_time:.0f}ms): {nombre} {apellido} como '{request.estado}'")

            return {
                "success": True,
                "mensaje": f"{nombre} {apellido} registrado como '{request.estado}'",
                "nuevo": True
            }

    except HTTPException:
        raise
    except Exception as error:
        response_time = (datetime.now() - start_time).total_seconds() * 1000
        logger.error(f"❌ Error ({response_time:.0f}ms): {error}")
        raise HTTPException(status_code=500, detail="Error del servidor")


@router.get("/resumen")
async def obtener_resumen(
    turno: str = Query("matutino"),
    connection: aiomysql.Connection = Depends(get_db_connection)
):
    """Obtener resumen de asistencias del día por turno"""
    try:
        datos = obtener_fecha_hora_cdmx()
        fecha = datos["fecha"]  # <-- esto es un objeto date
        hoy = fecha.strftime('%Y-%m-%d')
        sufijo_turno = 'M' if turno.lower() == 'matutino' else 'V'

        # Construimos patrón seguro para LIKE
        like_turno = f"%{sufijo_turno}%"
        
        async with connection.cursor() as cursor:
            # Total alumnos en el turno
            await cursor.execute(
                "SELECT COUNT(*) FROM estudiante e JOIN grupo g ON e.id_grupo = g.id_grupo WHERE g.nombre LIKE %s",
                (like_turno,)
            )
            total_alumnos = (await cursor.fetchone())[0]

            # Alumnos presentes hoy
            await cursor.execute(
                """
                SELECT COUNT(DISTINCT a.id_estudiante) FROM asistencia a
                JOIN estudiante e ON a.id_estudiante = e.id_estudiante
                JOIN grupo g ON e.id_grupo = g.id_grupo
                WHERE a.estado = 'presente'
                    AND a.fecha = %s
                    AND g.nombre LIKE %s
                """,
                (hoy, like_turno)
            )
            presentes = (await cursor.fetchone())[0]

            # Alumnos con justificante hoy
            await cursor.execute(
                """
                SELECT COUNT(DISTINCT a.id_estudiante) FROM asistencia a
                JOIN estudiante e ON a.id_estudiante = e.id_estudiante
                JOIN grupo g ON e.id_grupo = g.id_grupo
                WHERE a.estado = 'justificante'
                    AND a.fecha = %s
                    AND g.nombre LIKE %s
                """,
                (hoy, like_turno)
            )
            justificantes = (await cursor.fetchone())[0]

            ausentes = max(0, total_alumnos - presentes - justificantes)
            porcentaje = round((presentes / total_alumnos) * 100) if total_alumnos > 0 else 0

            return {
                "totalAlumnos": total_alumnos,
                "presentes": presentes,
                "justificantes": justificantes,
                "ausentes": ausentes,
                "porcentaje": porcentaje
            }

    except Exception as error:
        logger.error(f"❌ Error al obtener resumen del día: {error}")
        raise HTTPException(status_code=500, detail="Error al obtener resumen de asistencia diaria")
    
@router.get("/por-clase")
async def obtener_por_clase(fecha: Optional[str] = Query(None)):
    try:
        datos_fecha = obtener_fecha_hora_cdmx()
        hoy = datos_fecha["fecha"]
        dia_semana_texto = datos_fecha["dia"]

        if fecha:
            hoy = fecha
            dia_semana_texto = obtener_dia_semana_texto(hoy)

        logger.info(f"🔍 Consultando asistencia por clase - Fecha: {hoy}, Día: {dia_semana_texto}")

        query = """
        SELECT 
            c.id_clase,
            c.nombre_clase,
            g.nombre AS grupo,
            COUNT(CASE WHEN a.estado = 'presente' THEN 1 END) AS presentes,
            COUNT(CASE WHEN a.estado = 'justificante' THEN 1 END) AS justificantes,
            COUNT(DISTINCT e.id_estudiante) 
                - COUNT(CASE WHEN a.estado IN ('presente', 'justificante') THEN 1 END) AS ausentes
        FROM clase c
        JOIN grupo g ON g.id_grupo = c.id_grupo
        JOIN estudiante e ON e.id_grupo = g.id_grupo
        JOIN (
            SELECT DISTINCT id_clase
            FROM horario_clase
            WHERE dia = %s
        ) hc ON hc.id_clase = c.id_clase
        LEFT JOIN asistencia a 
            ON a.id_estudiante = e.id_estudiante 
            AND a.id_clase = c.id_clase 
            AND a.fecha = %s
        GROUP BY c.id_clase, c.nombre_clase, g.nombre
        ORDER BY c.id_clase
        """

        # Ejecutar query
        result = await fetch_all(query, (dia_semana_texto, hoy))
        
        # 🔧 Verificar si result es None o está vacío
        if result is None:
            logger.warning("⚠️ Query devolvió None")
            return []
        
        if not result:
            logger.info("ℹ️ No se encontraron clases para el día especificado")
            return []

        # 🔧 Verificar tipo de datos
        logger.info(f"📊 Total de registros: {len(result)}")
        if result:
            logger.info(f"📊 Tipo del primer elemento: {type(result[0])}")
            logger.info(f"📊 Primer elemento: {result[0]}")

        # Agregar porcentaje - con manejo de errores mejorado
        processed_result = []
        for i, row in enumerate(result):
            try:
                # 🔧 Verificar si row es dict o tuple
                if isinstance(row, dict):
                    total = row["presentes"] + row["justificantes"] + row["ausentes"]
                    row["porcentaje"] = round((row["presentes"] / total) * 100) if total > 0 else 0
                    processed_result.append(row)
                else:
                    # Si es tuple, convertir manualmente
                    logger.warning(f"⚠️ Fila {i} es tuple, convirtiendo a dict")
                    row_dict = {
                        "id_clase": row[0],
                        "nombre_clase": row[1], 
                        "grupo": row[2],
                        "presentes": row[3],
                        "justificantes": row[4],
                        "ausentes": row[5]
                    }
                    total = row_dict["presentes"] + row_dict["justificantes"] + row_dict["ausentes"]
                    row_dict["porcentaje"] = round((row_dict["presentes"] / total) * 100) if total > 0 else 0
                    processed_result.append(row_dict)
            except Exception as row_error:
                logger.error(f"❌ Error procesando fila {i}: {row_error}")
                logger.error(f"❌ Contenido de la fila: {row}")
                continue

        return processed_result

    except Exception as error:
        logger.error(f"❌ Error en /asistencia/por-clase: {error}")
        logger.error(f"❌ Tipo de error: {type(error)}")
        import traceback
        logger.error(f"❌ Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error al obtener resumen por clase: {str(error)}")

@router.get("/alumnos/clase/{id_clase}/excel")
async def generar_excel_clase(id_clase: int):
    """Generar Excel con asistencias de una clase específica"""
    try:
        # Obtener info del grupo y materia
        clase_info = await fetch_one("""
            SELECT 
                c.id_grupo,
                c.id_materia,
                g.nombre AS nombre_grupo,
                m.nombre AS nombre_materia
            FROM clase c
            JOIN grupo g ON c.id_grupo = g.id_grupo
            JOIN materia m ON c.id_materia = m.id_materia
            WHERE c.id_clase = %s
        """, (id_clase,))
        
        if not clase_info:
            raise HTTPException(status_code=404, detail="Clase no encontrada")
        
        id_grupo = clase_info['id_grupo']
        id_materia = clase_info['id_materia']
        nombre_grupo = clase_info['nombre_grupo']
        nombre_materia = clase_info['nombre_materia']
        
        # Obtener asistencias de todas las clases del mismo grupo y materia
        resultado = await fetch_all("""
            SELECT 
                e.id_estudiante,
                CONCAT(e.nombre, ' ', e.apellido) AS nombre,
                e.matricula,
                a.estado,
                a.fecha
            FROM estudiante e
            JOIN clase c ON c.id_grupo = e.id_grupo
            LEFT JOIN asistencia a ON a.id_estudiante = e.id_estudiante AND a.id_clase = c.id_clase
            WHERE c.id_grupo = %s AND c.id_materia = %s
            ORDER BY e.nombre, e.apellido, a.fecha
        """, (id_grupo, id_materia))
        
        # Resto del código de generación de Excel se queda igual...
        fechas_unicas = sorted(list(set(
            r['fecha'].strftime('%Y-%m-%d') for r in resultado 
            if r['fecha'] is not None
        )))
        
        estudiantes_map = {}
        for row in resultado:
            id_estudiante = row['id_estudiante']
            if id_estudiante not in estudiantes_map:
                estudiantes_map[id_estudiante] = {
                    'nombre': row['nombre'],
                    'matricula': row['matricula'],
                    'estados_por_fecha': {}
                }
            if row['fecha']:
                fecha_str = row['fecha'].strftime('%Y-%m-%d')
                estudiantes_map[id_estudiante]['estados_por_fecha'][fecha_str] = row['estado']
        
        # Crear workbook y worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Asistencias'
        
        encabezados = ['Nombre', 'Matrícula', 'Grupo', 'Materia'] + fechas_unicas
        worksheet.append(encabezados)
        
        def estado_to_letra_color(estado):
            if not estado:
                return {'letra': 'A', 'color': 'FFFF0000'}
            estado_lower = estado.lower()
            return {
                'presente': {'letra':'P','color':'FF00B050'},
                'ausente': {'letra':'A','color':'FFFF0000'},
                'justificante': {'letra':'J','color':'FFFFC000'}
            }.get(estado_lower, {'letra':'A','color':'FFFF0000'})
        
        for estudiante_data in estudiantes_map.values():
            fila = [
                estudiante_data['nombre'],
                estudiante_data['matricula'],
                nombre_grupo,
                nombre_materia
            ]
            for fecha in fechas_unicas:
                estado = estudiante_data['estados_por_fecha'].get(fecha)
                letra = estado_to_letra_color(estado)['letra']
                fila.append(letra)
            worksheet.append(fila)
        
        # Formato
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for col_idx, fecha in enumerate(fechas_unicas, 5):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                estado_letra = cell.value
                color_map = {'P':'FF00B050','A':'FFFF0000','J':'FFFFC000'}
                if estado_letra in color_map:
                    cell.fill = PatternFill(start_color=color_map[estado_letra], 
                                            end_color=color_map[estado_letra], 
                                            fill_type='solid')
                    cell.font = Font(color='FFFFFFFF', bold=True)
                    cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho columnas
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 25
        for col_idx in range(5, 5 + len(fechas_unicas)):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 12
        
        # Guardar en memoria
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=asistencias.xlsx'}
        )
    
    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"❌ Error al generar Excel: {error}")
        raise HTTPException(status_code=500, detail="Error al generar Excel")

@router.post("/actualizar")
async def actualizar_asistencia(request: ActualizarAsistenciaRequest):
    """Actualizar estado de asistencia por matrícula"""
    try:
        fecha = obtener_fecha_hora_cdmx()['fecha']

        # Buscar al estudiante
        estudiante_result = await fetch_one(
            "SELECT id_estudiante FROM estudiante WHERE matricula = %s",
            (request.matricula,)
        )
        if not estudiante_result:
            raise HTTPException(status_code=404, detail="Estudiante no encontrado")
        
        id_estudiante = estudiante_result['id_estudiante']

        # Verificar si ya existe registro
        asistencia_result = await fetch_one(
            "SELECT id_asistencia FROM asistencia WHERE id_estudiante = %s AND id_clase = %s AND fecha = %s",
            (id_estudiante, request.id_clase, fecha)
        )

        if asistencia_result:
            # Actualizar registro existente
            await execute_query(
                "UPDATE asistencia SET estado = %s WHERE id_estudiante = %s AND id_clase = %s AND fecha = %s",
                (request.estado, id_estudiante, request.id_clase, fecha)
            )
        else:
            # Insertar nuevo registro
            await execute_query(
                "INSERT INTO asistencia (id_estudiante, id_clase, estado, fecha) VALUES (%s, %s, %s, %s)",
                (id_estudiante, request.id_clase, request.estado, fecha)
            )

        return {"success": True, "mensaje": "Estado actualizado"}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"Error actualizando asistencia: {error}")
        raise HTTPException(status_code=500, detail="Error actualizando asistencia")
    
@router.put("/estado")
async def actualizar_estado(request: ActualizarEstadoRequest):
    """Actualizar estado de asistencia por ID de estudiante"""
    try:
        fecha_hora = obtener_fecha_hora_cdmx()
        fecha = fecha_hora['fecha']
        hora = fecha_hora['hora']

        # Verificar si existe registro
        existing = await fetch_one(
            "SELECT * FROM asistencia WHERE id_estudiante = %s AND id_clase = %s AND fecha = %s",
            (request.id_estudiante, request.id_clase, fecha)
        )

        if existing:
            # Actualizar existente
            hora_entrada = hora if request.estado == 'presente' else existing.get('hora_entrada')
            await execute_query(
                """
                UPDATE asistencia
                SET estado = %s,
                    hora_entrada = %s
                WHERE id_estudiante = %s AND id_clase = %s AND fecha = %s
                """,
                (request.estado, hora_entrada, request.id_estudiante, request.id_clase, fecha)
            )
            return {"message": "Estado de asistencia actualizado"}

        # Insertar nuevo registro
        hora_entrada = hora if request.estado == 'presente' else None
        await execute_query(
            "INSERT INTO asistencia (id_estudiante, id_clase, estado, hora_entrada, fecha) VALUES (%s, %s, %s, %s, %s)",
            (request.id_estudiante, request.id_clase, request.estado, hora_entrada, fecha)
        )

        return {"message": "Asistencia registrada correctamente"}

    except Exception as error:
        logger.error(f"Error al registrar o actualizar asistencia: {error}")
        raise HTTPException(status_code=500, detail="Error al registrar asistencia")

@router.put("/actualizar-estado")
async def actualizar_estado_alt(request: ActualizarEstadoRequest):
    """Endpoint alternativo para actualizar estado de asistencia"""
    try:
        fecha_hora = obtener_fecha_hora_cdmx()
        hoy = fecha_hora['fecha']
        hora = fecha_hora['hora']

        # Verificar que existe el registro
        existing = await fetch_one(
            "SELECT * FROM asistencia WHERE id_estudiante = %s AND id_clase = %s AND fecha = %s",
            (request.id_estudiante, request.id_clase, hoy)
        )

        if not existing:
            raise HTTPException(status_code=404, detail="Registro de asistencia no encontrado para actualizar")

        # Actualizar estado
        hora_entrada = hora if request.estado.lower() == 'presente' else existing.get('hora_entrada')
        await execute_query(
            """
            UPDATE asistencia
            SET estado = %s,
                hora_entrada = %s
            WHERE id_estudiante = %s AND id_clase = %s AND fecha = %s
            """,
            (request.estado, hora_entrada, request.id_estudiante, request.id_clase, hoy)
        )

        return {"message": "Estado actualizado correctamente"}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"Error al actualizar estado: {error}")
        raise HTTPException(status_code=500, detail="Error interno al actualizar estado")

@router.get("/resumen-general")
async def obtener_resumen_general(
    turno: Optional[str] = Query(None),
    connection: aiomysql.Connection = Depends(get_db_connection)
):
    try:
        fecha_hora = obtener_fecha_hora_cdmx()
        hoy = fecha_hora['fecha']
        dia = fecha_hora['dia']

        hora_inicio_turno = "07:00:00"
        hora_fin_turno = "13:30:00"
        if turno == "vespertino":
            hora_inicio_turno = "13:35:00"
            hora_fin_turno = "19:20:00"

        async with connection.cursor(aiomysql.DictCursor) as cursor:
            # Total alumnos
            await cursor.execute("SELECT COUNT(*) AS total FROM estudiante")
            total_result = await cursor.fetchone()
            total_alumnos = total_result['total'] if total_result else 0

            # Contar ausentes, presentes y justificantes
            query = """
                SELECT
                    COUNT(DISTINCT CASE WHEN a.estado = 'ausente' THEN a.id_estudiante END) AS ausentes,
                    COUNT(DISTINCT CASE WHEN a.estado = 'presente' THEN a.id_estudiante END) AS presentes,
                    COUNT(DISTINCT CASE WHEN a.estado = 'justificante' THEN a.id_estudiante END) AS justificantes
                FROM asistencia a
                JOIN clase c ON a.id_clase = c.id_clase
                JOIN horario_clase hc ON hc.id_clase = c.id_clase
                WHERE a.fecha BETWEEN %s AND %s
                  AND hc.hora_inicio >= %s
                  AND hc.hora_inicio <= %s
                  AND LOWER(hc.dia) = %s
            """
            await cursor.execute(query, (FECHA_INICIO_CICLO, hoy, hora_inicio_turno, hora_fin_turno, dia.lower()))
            result = await cursor.fetchone()

        ausentes_num = int(result['ausentes']) if result and result['ausentes'] else 0
        presentes_num = int(result['presentes']) if result and result['presentes'] else 0
        justificantes_num = int(result['justificantes']) if result and result['justificantes'] else 0

        porcentaje = round((presentes_num / total_alumnos) * 100) if total_alumnos > 0 else 0

        return {
            "totalAlumnos": total_alumnos,
            "presentes": presentes_num,
            "justificantes": justificantes_num,
            "ausentes": ausentes_num,
            "porcentaje": porcentaje
        }

    except Exception as error:
        logger.error(f"❌ Error al obtener resumen general: {error}")
        raise HTTPException(status_code=500, detail="Error al obtener resumen general")

@router.get("/lista-alumnos")
async def obtener_lista_alumnos(
    turno: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    connection: aiomysql.Connection = Depends(get_db_connection)
):
    """Obtener lista de alumnos filtrada por turno y estado"""
    try:
        fecha_hora = obtener_fecha_hora_cdmx()
        hoy = fecha_hora['fecha']
        dia = fecha_hora['dia']
        dia_semana_texto = dia.lower()
        
        hora_inicio_turno = "07:00:00"
        hora_fin_turno = "13:30:00"
        if turno == "vespertino":
            hora_inicio_turno = "13:35:00"
            hora_fin_turno = "19:20:00"
        
        async with connection.cursor(aiomysql.DictCursor) as cursor:
            await cursor.execute("""
                SELECT
                    e.nombre,
                    e.apellido,
                    e.matricula,
                    g.nombre AS grupo,
                    COUNT(a.id_asistencia) AS totalFaltas
                FROM asistencia a
                JOIN estudiante e ON e.id_estudiante = a.id_estudiante
                JOIN grupo g ON g.id_grupo = e.id_grupo
                JOIN clase c ON a.id_clase = c.id_clase
                JOIN horario_clase hc ON hc.id_clase = c.id_clase
                WHERE a.fecha BETWEEN %s AND %s
                    AND a.estado = %s
                    AND LOWER(hc.dia) = %s
                    AND hc.hora_inicio >= %s
                    AND hc.hora_inicio <= %s
                GROUP BY e.nombre, e.apellido, e.matricula, g.nombre
                ORDER BY totalFaltas DESC
            """, (FECHA_INICIO_CICLO, hoy, estado, dia_semana_texto, hora_inicio_turno, hora_fin_turno))
            
            result = await cursor.fetchall()
            return result
            
    except Exception as error:
        logger.error(f"❌ Error en lista-alumnos: {error}")
        raise HTTPException(status_code=500, detail="Error al obtener lista de alumnos")
    
@router.get("/excel-general")
async def generar_excel_general(
    turno: Optional[str] = Query(None),
    connection: aiomysql.Connection = Depends(get_db_connection)
):
    """Generar Excel con resumen general acumulado"""
    try:
        fecha_hora = obtener_fecha_hora_cdmx()
        hoy = fecha_hora['fecha']
        dia = fecha_hora['dia']
        dia_semana_texto = dia.lower()
        
        hora_inicio_turno = "07:00:00"
        hora_fin_turno = "13:30:00"
        if turno == "vespertino":
            hora_inicio_turno = "13:35:00"
            hora_fin_turno = "19:20:00"
        
        async with connection.cursor(aiomysql.DictCursor) as cursor:
            await cursor.execute("""
                SELECT
                    e.nombre,
                    e.apellido,
                    e.matricula,
                    g.nombre AS grupo,
                    SUM(CASE WHEN a.estado = 'presente' THEN 1 ELSE 0 END) AS presentes,
                    SUM(CASE WHEN a.estado = 'ausente' THEN 1 ELSE 0 END) AS ausentes,
                    SUM(CASE WHEN a.estado = 'justificante' THEN 1 ELSE 0 END) AS justificantes
                FROM asistencia a
                JOIN estudiante e ON e.id_estudiante = a.id_estudiante
                JOIN grupo g ON g.id_grupo = e.id_grupo
                JOIN clase c ON a.id_clase = c.id_clase
                JOIN horario_clase hc ON hc.id_clase = c.id_clase
                WHERE a.fecha BETWEEN %s AND %s
                    AND LOWER(hc.dia) = %s
                    AND hc.hora_inicio >= %s
                    AND hc.hora_inicio <= %s
                GROUP BY e.nombre, e.apellido, e.matricula, g.nombre
                ORDER BY e.apellido, e.nombre
            """, (FECHA_INICIO_CICLO, hoy, dia_semana_texto, hora_inicio_turno, hora_fin_turno))
            
            result = await cursor.fetchall()
        
        # Crear workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Asistencia General"
        
        # Definir encabezados
        encabezados = ["Nombre", "Apellido", "Matrícula", "Grupo", "Presentes", "Ausentes", "Justificantes"]
        worksheet.append(encabezados)
        
        # Agregar datos
        for row in result:
            worksheet.append([
                row['nombre'],
                row['apellido'],
                row['matricula'],
                row['grupo'],
                row['presentes'],
                row['ausentes'],
                row['justificantes']
            ])
        
        # Aplicar formato a encabezados
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        
        # Ajustar ancho de columnas
        column_widths = [20, 20, 15, 10, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            worksheet.column_dimensions[col_letter].width = width
        
        # Guardar en memoria
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=asistencia_{turno}.xlsx'}
        )
        
    except Exception as error:
        logger.error(f"❌ Error en excel-general: {error}")
        raise HTTPException(status_code=500, detail="Error al generar excel")
    
# ✅ Función auxiliar para inicializar asistencias de una clase
async def inicializar_asistencias(id_clase: int):
    """
    Inserta asistencias 'ausente' para todos los estudiantes de la clase
    si aún no existen registros para la fecha actual.
    """
    datos_fecha = obtener_fecha_hora_cdmx()
    hoy = datos_fecha["fecha"]

    # 1️⃣ Obtener grupo de la clase
    query_grupo = "SELECT id_grupo FROM clase WHERE id_clase = %s"
    grupo = await fetch_all(query_grupo, (id_clase,))
    if not grupo:
        return
    id_grupo = grupo[0]["id_grupo"]

    # 2️⃣ Obtener estudiantes del grupo
    query_estudiantes = "SELECT id_estudiante FROM estudiante WHERE id_grupo = %s"
    estudiantes = await fetch_all(query_estudiantes, (id_grupo,))

    # 3️⃣ Insertar asistencia como 'ausente' si no existe
    for est in estudiantes:
        query_insert = """
            INSERT INTO asistencia (id_estudiante, id_clase, fecha, estado)
            SELECT %s, %s, %s, %s
            WHERE NOT EXISTS (
                SELECT 1 FROM asistencia 
                WHERE id_estudiante = %s AND id_clase = %s AND fecha = %s
            )
        """
        await execute_query(query_insert, (
            est["id_estudiante"], id_clase, hoy, "ausente",
            est["id_estudiante"], id_clase, hoy
        ))

# ✅ Endpoint: /api/asistencia/clase/{id_clase}
@router.get("/clase/{id_clase}")
async def obtener_asistencia_clase(id_clase: int):
    try:
        # Inicializar asistencias
        await inicializar_asistencias(id_clase)

        # Fecha local en CDMX
        datos_fecha = obtener_fecha_hora_cdmx()
        hoy = datos_fecha["fecha"]

        # Consultar asistencia de los estudiantes
        query = """
            SELECT 
                e.id_estudiante,
                e.matricula,
                e.nombre,
                e.apellido,
                e.no_lista,
                COALESCE(a.estado, 'ausente') AS estado
            FROM estudiante e
            LEFT JOIN asistencia a 
                ON e.id_estudiante = a.id_estudiante 
                AND a.id_clase = %s 
                AND a.fecha = %s
            WHERE e.id_grupo = (
                SELECT id_grupo FROM clase WHERE id_clase = %s
            )
            ORDER BY e.no_lista
        """
        rows = await fetch_all(query, (id_clase, hoy, id_clase))

        return rows

    except Exception as e:
        print("❌ Error en /api/asistencia/clase/{id_clase}:", str(e))
        raise HTTPException(status_code=500, detail="Error al obtener asistencia de la clase")  
