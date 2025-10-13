from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timedelta
import io
from typing import Optional
from config.db import fetch_all, fetch_one
from utils.fecha import obtener_fecha_hora_cdmx
import traceback

router = APIRouter()

# Colores para estados
COLOR_PRESENTE = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
COLOR_AUSENTE = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
COLOR_JUSTIFICANTE = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
COLOR_ENTREGADO = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
COLOR_PENDIENTE = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
COLOR_NO_ENTREGADO = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

def generar_rango_fechas(fecha_inicio: str, fecha_fin: str):
    """Genera lista de fechas entre inicio y fin"""
    inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    fin = datetime.strptime(fecha_fin, "%Y-%m-%d")
    fechas = []
    
    while inicio <= fin:
        fechas.append(inicio.strftime("%Y-%m-%d"))
        inicio += timedelta(days=1)
    
    return fechas

def limpiar_nombre_hoja(nombre: str, max_length: int = 31) -> str:
    """Limpia el nombre de la hoja para Excel"""
    invalidos = ['*', '?', ':', '\\', '/', '[', ']']
    for char in invalidos:
        nombre = nombre.replace(char, '_')
    return nombre[:max_length]

def formato_fecha(fecha):
    """Convierte fecha a string YYYY-MM-DD"""
    if isinstance(fecha, datetime):
        return fecha.strftime("%Y-%m-%d")
    elif isinstance(fecha, str):
        return fecha
    return str(fecha)

# ==================== REPORTE DE ASISTENCIAS POR GRUPO ====================
@router.get("/excel")
async def generar_reporte_grupo(
    id_grupo: int = Query(..., description="ID del grupo"),
    fechaInicio: Optional[str] = Query(None, description="Fecha inicio YYYY-MM-DD"),
    fechaFin: Optional[str] = Query(None, description="Fecha fin YYYY-MM-DD")
):
    """Genera reporte Excel de asistencias por grupo con filtros de fecha"""
    try:
        # Si no hay fechas, usar valores por defecto
        if not fechaInicio:
            min_fecha = await fetch_one("SELECT MIN(fecha) as min_fecha FROM asistencia")
            fechaInicio = formato_fecha(min_fecha["min_fecha"]) if min_fecha and min_fecha["min_fecha"] else "2025-08-04"
        
        if not fechaFin:
            fechaFin = obtener_fecha_hora_cdmx()["fecha"]
        
        print(f"📅 Rango usado: {fechaInicio} a {fechaFin}")
        
        # Generar lista de fechas
        fechas = generar_rango_fechas(fechaInicio, fechaFin)
        
        # ✅ CONSULTA CORREGIDA: Obtener todas las clases del grupo con sus estudiantes
        query = """
            SELECT
                m.nombre AS materia,
                e.nombre,
                e.apellido,
                e.matricula,
                e.no_lista,
                a.fecha,
                a.estado,
                c.id_clase
            FROM estudiante e
            CROSS JOIN clase c
            LEFT JOIN materia m ON c.id_materia = m.id_materia
            LEFT JOIN asistencia a 
                ON a.id_estudiante = e.id_estudiante
                AND a.id_clase = c.id_clase
                AND a.fecha BETWEEN %s AND %s
            WHERE e.id_grupo = %s
                AND c.id_grupo = %s
            ORDER BY m.nombre, e.no_lista, a.fecha
        """
        
        datos = await fetch_all(query, (fechaInicio, fechaFin, id_grupo, id_grupo))
        
        if not datos:
            raise HTTPException(status_code=404, detail="No se encontraron datos para este grupo")
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Agrupar por materia
        materias = {}
        for row in datos:
            mat = row["materia"]
            if mat not in materias:
                materias[mat] = []
            materias[mat].append(row)
        
        # Crear hoja por materia
        for materia, rows in materias.items():
            ws = wb.create_sheet(limpiar_nombre_hoja(materia))
            
            # Encabezados
            headers = ["No. Lista", "Nombre", "Apellido", "Matrícula", *fechas]
            ws.append(headers)
            
            # Aplicar estilo al encabezado
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Agrupar por alumno
            alumnos = {}
            for row in rows:
                key = row["matricula"]
                if key not in alumnos:
                    alumnos[key] = {
                        "nombre": row["nombre"],
                        "apellido": row["apellido"],
                        "matricula": row["matricula"],
                        "no_lista": row["no_lista"],
                        "asistencias": {}
                    }
                if row["fecha"]:
                    fecha_str = formato_fecha(row["fecha"])
                    alumnos[key]["asistencias"][fecha_str] = row["estado"]
            
            # Ordenar alumnos por no_lista
            alumnos_ordenados = sorted(alumnos.values(), key=lambda x: x["no_lista"])
            
            # Agregar filas
            for alumno in alumnos_ordenados:
                row_data = [
                    alumno["no_lista"],
                    alumno["nombre"],
                    alumno["apellido"],
                    alumno["matricula"]
                ]
                
                # Agregar estados por fecha
                for fecha in fechas:
                    estado = alumno["asistencias"].get(fecha, "")
                    row_data.append(estado.capitalize() if estado else "")
                
                ws.append(row_data)
                
                # Colorear celdas de estado
                current_row = ws.max_row
                for col_idx, fecha in enumerate(fechas, start=5):
                    cell = ws.cell(row=current_row, column=col_idx)
                    estado = cell.value.lower() if cell.value else ""
                    
                    if estado == "presente":
                        cell.fill = COLOR_PRESENTE
                        cell.alignment = Alignment(horizontal="center")
                    elif estado == "ausente":
                        cell.fill = COLOR_AUSENTE
                        cell.alignment = Alignment(horizontal="center")
                    elif estado == "justificante":
                        cell.fill = COLOR_JUSTIFICANTE
                        cell.alignment = Alignment(horizontal="center")
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            for i in range(5, 5 + len(fechas)):
                ws.column_dimensions[chr(64 + i)].width = 12
        
        # Guardar en buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=asistencias_grupo_{id_grupo}.xlsx"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")

# ==================== REPORTE INDIVIDUAL DE ESTUDIANTE ====================
@router.get("/excel/individual")
async def generar_reporte_individual(
    id_estudiante: int = Query(..., description="ID del estudiante"),
    fechaInicio: str = Query(..., description="Fecha inicio YYYY-MM-DD"),
    fechaFin: str = Query(..., description="Fecha fin YYYY-MM-DD")
):
    """Genera reporte individual de un estudiante"""
    try:
        print(f"📅 Rango usado: {fechaInicio} a {fechaFin}")
        
        # Obtener datos del estudiante
        estudiante = await fetch_one(
            "SELECT nombre, apellido, matricula FROM estudiante WHERE id_estudiante = %s",
            (id_estudiante,)
        )
        
        if not estudiante:
            raise HTTPException(status_code=404, detail="Estudiante no encontrado")
        
        # Generar fechas
        fechas = generar_rango_fechas(fechaInicio, fechaFin)
        
        # ✅ CONSULTA CORREGIDA: Obtener asistencias agrupadas por materia
        query = """
            SELECT
                m.nombre AS materia,
                a.fecha,
                a.estado
            FROM asistencia a
            JOIN clase c ON a.id_clase = c.id_clase
            JOIN materia m ON c.id_materia = m.id_materia
            WHERE a.id_estudiante = %s
                AND a.fecha BETWEEN %s AND %s
            ORDER BY m.nombre, a.fecha
        """
        
        asistencias = await fetch_all(query, (id_estudiante, fechaInicio, fechaFin))
        
        # Agrupar por materia
        materias = {}
        for row in asistencias:
            mat = row["materia"]
            if mat not in materias:
                materias[mat] = {}
            fecha_str = formato_fecha(row["fecha"])
            materias[mat][fecha_str] = row["estado"]
        
        if not materias:
            raise HTTPException(status_code=404, detail="No se encontraron asistencias para este estudiante en el rango de fechas")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencias"
        
        # Encabezados
        headers = ["Materia", *fechas]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar filas por materia
        for materia, asist_fechas in sorted(materias.items()):
            row_data = [materia]
            
            for fecha in fechas:
                estado = asist_fechas.get(fecha, "")
                row_data.append(estado.capitalize() if estado else "")
            
            ws.append(row_data)
            
            # Colorear celdas
            current_row = ws.max_row
            for col_idx in range(2, len(fechas) + 2):
                cell = ws.cell(row=current_row, column=col_idx)
                estado = cell.value.lower() if cell.value else ""
                
                if estado == "presente":
                    cell.fill = COLOR_PRESENTE
                    cell.alignment = Alignment(horizontal="center")
                elif estado == "ausente":
                    cell.fill = COLOR_AUSENTE
                    cell.alignment = Alignment(horizontal="center")
                elif estado == "justificante":
                    cell.fill = COLOR_JUSTIFICANTE
                    cell.alignment = Alignment(horizontal="center")
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 25
        for i in range(2, 2 + len(fechas)):
            ws.column_dimensions[chr(64 + i)].width = 12
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"reporte_alumno_{estudiante['matricula']}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")

# ==================== REPORTE DE ACTIVIDADES POR CLASE ====================
@router.get("/excel/clase/{id_clase}")
async def generar_reporte_actividades_clase(id_clase: int):
    """Genera reporte Excel con actividades de una clase (una hoja por actividad)"""
    try:
        # 1. Obtener datos de la clase
        clase = await fetch_one("""
            SELECT 
                c.id_clase,
                m.nombre AS materia,
                g.nombre AS grupo
            FROM clase c
            LEFT JOIN materia m ON c.id_materia = m.id_materia
            LEFT JOIN grupo g ON c.id_grupo = g.id_grupo
            WHERE c.id_clase = %s
        """, (id_clase,))
        
        if not clase:
            raise HTTPException(status_code=404, detail="Clase no encontrada")
        
        # 2. Obtener actividades de la clase
        actividades = await fetch_all("""
            SELECT id_actividad, titulo, fecha_entrega, valor_maximo
            FROM actividad
            WHERE id_clase = %s
            ORDER BY fecha_entrega ASC
        """, (id_clase,))
        
        if not actividades:
            raise HTTPException(status_code=404, detail="No hay actividades para esta clase")
        
        # 3. Obtener alumnos del grupo
        alumnos = await fetch_all("""
            SELECT id_estudiante, nombre, apellido, matricula
            FROM estudiante
            WHERE id_grupo = (
                SELECT id_grupo FROM clase WHERE id_clase = %s
            )
            ORDER BY apellido, nombre
        """, (id_clase,))
        
        # 4. Crear workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # 5. Crear una hoja por cada actividad
        for act in actividades:
            ws = wb.create_sheet(limpiar_nombre_hoja(act["titulo"]))
            
            # Información de la actividad
            ws.append([f"Actividad: {act['titulo']}"])
            ws.append([f"Fecha entrega: {formato_fecha(act['fecha_entrega'])}"])
            ws.append([f"Valor máximo: {act['valor_maximo']}"])
            ws.append([])
            
            # Encabezados
            ws.append(["Alumno", "Matrícula", "Estado", "Fecha entrega real", "Calificación"])
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # 6. Obtener entregas de la actividad
            entregas = await fetch_all("""
                SELECT id_estudiante, estado, fecha_entrega_real, calificacion
                FROM actividad_estudiante
                WHERE id_actividad = %s
            """, (act["id_actividad"],))
            
            entregas_map = {e["id_estudiante"]: e for e in entregas}
            
            # 7. Agregar fila por cada alumno
            for alumno in alumnos:
                entrega = entregas_map.get(alumno["id_estudiante"])
                estado = entrega["estado"] if entrega else "pendiente"
                
                row_data = [
                    f"{alumno['apellido']} {alumno['nombre']}",
                    alumno["matricula"],
                    estado.capitalize(),
                    formato_fecha(entrega["fecha_entrega_real"]) if entrega and entrega.get("fecha_entrega_real") else "",
                    entrega["calificacion"] if entrega and entrega.get("calificacion") is not None else ""
                ]
                
                ws.append(row_data)
                
                # 🎨 Colorear celda de estado
                current_row = ws.max_row
                estado_cell = ws.cell(row=current_row, column=3)
                estado_cell.alignment = Alignment(horizontal="center")
                
                estado_lower = estado.lower()
                if estado_lower == "entregado":
                    estado_cell.fill = COLOR_ENTREGADO
                elif estado_lower == "pendiente":
                    estado_cell.fill = COLOR_PENDIENTE
                elif estado_lower == "no entregado":
                    estado_cell.fill = COLOR_NO_ENTREGADO
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
        
        # 8. Generar archivo y enviarlo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Actividades_{clase['materia']}_{clase['grupo']}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error generando Excel: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generando el Excel: {str(e)}")

# ==================== REPORTE GENERAL DE ACTIVIDADES ====================
@router.get("/excel/clase/general/{id_clase}")
async def generar_reporte_general_actividades(id_clase: int):
    """Genera reporte general con todas las actividades en una sola hoja"""
    try:
        # Obtener clase
        clase = await fetch_one("""
            SELECT c.id_clase, m.nombre AS materia, g.nombre AS grupo
            FROM clase c
            LEFT JOIN materia m ON c.id_materia = m.id_materia
            LEFT JOIN grupo g ON c.id_grupo = g.id_grupo
            WHERE c.id_clase = %s
        """, (id_clase,))
        
        if not clase:
            raise HTTPException(status_code=404, detail="Clase no encontrada")
        
        # Obtener actividades ordenadas
        actividades = await fetch_all("""
            SELECT id_actividad, titulo
            FROM actividad
            WHERE id_clase = %s
            ORDER BY fecha_entrega ASC
        """, (id_clase,))
        
        if not actividades:
            raise HTTPException(status_code=404, detail="No hay actividades")
        
        # Obtener alumnos
        alumnos = await fetch_all("""
            SELECT id_estudiante, nombre, apellido, matricula, no_lista
            FROM estudiante
            WHERE id_grupo = (SELECT id_grupo FROM clase WHERE id_clase = %s)
            ORDER BY no_lista
        """, (id_clase,))
        
        # Obtener todas las entregas de las actividades
        ids_actividades = [a["id_actividad"] for a in actividades]
        placeholders = ','.join(['%s'] * len(ids_actividades))
        entregas = await fetch_all(f"""
            SELECT id_estudiante, id_actividad, estado, calificacion
            FROM actividad_estudiante
            WHERE id_actividad IN ({placeholders})
        """, tuple(ids_actividades))
        
        entregas_map = {}
        for e in entregas:
            key = f"{e['id_estudiante']}_{e['id_actividad']}"
            entregas_map[key] = e
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = limpiar_nombre_hoja(f"{clase['materia']}-{clase['grupo']}")
        
        # Encabezados
        headers = ["No Lista", "Matrícula", "Nombre", "Grupo"]
        for act in actividades:
            headers.append(f"{act['titulo']}")
            headers.append(f"Cal")
        
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar filas de alumnos
        for alumno in alumnos:
            row_data = [
                alumno["no_lista"],
                alumno["matricula"],
                f"{alumno['apellido']} {alumno['nombre']}",
                clase["grupo"]
            ]
            
            for act in actividades:
                key = f"{alumno['id_estudiante']}_{act['id_actividad']}"
                entrega = entregas_map.get(key)
                
                estado = entrega["estado"].capitalize() if entrega else "Pendiente"
                calificacion = entrega["calificacion"] if entrega and entrega["calificacion"] is not None else 0
                
                row_data.append(estado)
                row_data.append(calificacion)
            
            ws.append(row_data)
            
            # Colorear estados
            current_row = ws.max_row
            col_idx = 5  # Empieza después de No Lista, Matrícula, Nombre, Grupo
            for _ in actividades:
                estado_cell = ws.cell(row=current_row, column=col_idx)
                estado_cell.alignment = Alignment(horizontal="center")
                estado = estado_cell.value.lower() if estado_cell.value else ""
                
                if estado == "entregado":
                    estado_cell.fill = COLOR_ENTREGADO
                elif estado == "pendiente":
                    estado_cell.fill = COLOR_PENDIENTE
                elif estado == "no entregado":
                    estado_cell.fill = COLOR_NO_ENTREGADO
                
                # Centrar también la calificación
                cal_cell = ws.cell(row=current_row, column=col_idx + 1)
                cal_cell.alignment = Alignment(horizontal="center")
                
                col_idx += 2  # Siguiente par Estado/Calificación
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        
        col_letra = 'E'
        for _ in actividades:
            ws.column_dimensions[col_letra].width = 15
            col_letra = chr(ord(col_letra) + 1)
            ws.column_dimensions[col_letra].width = 8
            col_letra = chr(ord(col_letra) + 1)
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Actividades_General_{clase['materia']}_{clase['grupo']}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")

# ==================== REPORTE DE PROFESOR ====================
@router.get("/excel/profesor/{id_profesor}")
async def generar_reporte_profesor(
    id_profesor: int,
    fechaInicio: Optional[str] = Query("2025-08-04"),
    fechaFin: Optional[str] = Query(None)
):
    """Genera reporte de todas las clases de un profesor"""
    try:
        if not fechaFin:
            fechaFin = obtener_fecha_hora_cdmx()["fecha"]
        
        print(f"📅 Generando reporte para profesor {id_profesor}: {fechaInicio} a {fechaFin}")
        
        # Obtener clases del profesor
        clases = await fetch_all("""
            SELECT c.id_clase, c.nombre_clase,
                   m.nombre AS materia, g.nombre AS grupo, c.id_grupo
            FROM clase c
            LEFT JOIN materia m ON c.id_materia = m.id_materia
            LEFT JOIN grupo g ON c.id_grupo = g.id_grupo
            WHERE c.id_profesor = %s
            ORDER BY m.nombre, g.nombre
        """, (id_profesor,))
        
        if not clases:
            raise HTTPException(status_code=404, detail="El profesor no tiene clases asignadas")
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for clase in clases:
            nombre_hoja = f"{clase['materia']} - {clase['grupo']}"
            ws = wb.create_sheet(limpiar_nombre_hoja(nombre_hoja))
            
            # Obtener estudiantes del grupo
            estudiantes = await fetch_all("""
                SELECT e.id_estudiante, e.nombre, e.apellido, e.matricula, 
                       e.no_lista, g.nombre AS grupo
                FROM estudiante e
                JOIN grupo g ON e.id_grupo = g.id_grupo
                WHERE e.id_grupo = %s
                ORDER BY e.no_lista
            """, (clase["id_grupo"],))
            
            if not estudiantes:
                ws.append(["No hay estudiantes en este grupo"])
                continue
            
            # Obtener solo fechas donde hubo clase (registros de asistencia)
            fechas_clase = await fetch_all("""
                SELECT DISTINCT fecha
                FROM asistencia
                WHERE id_clase = %s
                    AND fecha BETWEEN %s AND %s
                ORDER BY fecha
            """, (clase["id_clase"], fechaInicio, fechaFin))
            
            fechas = [formato_fecha(f["fecha"]) for f in fechas_clase]
            
            if not fechas:
                ws.append(["No hay registros de asistencia en este rango de fechas"])
                continue
            
            # Obtener asistencias
            asistencias = await fetch_all("""
                SELECT id_estudiante, fecha, estado
                FROM asistencia
                WHERE id_clase = %s
                    AND fecha BETWEEN %s AND %s
            """, (clase["id_clase"], fechaInicio, fechaFin))
            
            asist_map = {}
            for a in asistencias:
                fecha_str = formato_fecha(a["fecha"])
                key = f"{a['id_estudiante']}_{fecha_str}"
                asist_map[key] = a["estado"]
            
            # Encabezados
            headers = ["No Lista", "Matrícula", "Nombre", "Grupo", *fechas]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Agregar filas de estudiantes
            for est in estudiantes:
                row_data = [
                    est["no_lista"],
                    est["matricula"],
                    f"{est['apellido']} {est['nombre']}",
                    est["grupo"]
                ]
                
                for fecha in fechas:
                    key = f"{est['id_estudiante']}_{fecha}"
                    estado = asist_map.get(key, "")
                    row_data.append(estado.capitalize() if estado else "")
                
                ws.append(row_data)
                
                # Colorear
                current_row = ws.max_row
                for col_idx in range(5, 5 + len(fechas)):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.alignment = Alignment(horizontal="center")
                    estado = cell.value.lower() if cell.value else ""
                    
                    if estado == "presente":
                        cell.fill = COLOR_PRESENTE
                    elif estado == "ausente":
                        cell.fill = COLOR_AUSENTE
                    elif estado == "justificante":
                        cell.fill = COLOR_JUSTIFICANTE
            
            # Ajustar columnas
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 10
            for i in range(5, 5 + len(fechas)):
                if i <= 26:
                    col_letra = chr(64 + i)
                else:
                    col_letra = chr(64 + (i // 26)) + chr(64 + (i % 26))
                ws.column_dimensions[col_letra].width = 12
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Reporte_Profesor_{id_profesor}.xlsx"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")
    
@router.get("/excel/clase/completo/{id_clase}")
async def generar_reporte_completo_clase(id_clase: int):
    """Genera reporte Excel completo con actividades y asistencias"""
    try:
        def nombre_hoja_seguro(base: str, sufijo: str, max_len: int = 31) -> str:
            nombre_base = base[:max_len - len(sufijo) - 1] if len(base) > (max_len - len(sufijo) - 1) else base
            return f"{nombre_base}-{sufijo}"

        # 1️⃣ Info de clase
        clase = await fetch_one("""
            SELECT c.id_clase, m.nombre AS materia, g.nombre AS grupo, c.id_grupo, c.id_materia
            FROM clase c
            LEFT JOIN materia m ON c.id_materia = m.id_materia
            LEFT JOIN grupo g ON c.id_grupo = g.id_grupo
            WHERE c.id_clase = %s
        """, (id_clase,))
        if not clase:
            raise HTTPException(status_code=404, detail="Clase no encontrada")

        materia = clase["materia"]
        grupo = clase["grupo"]
        id_grupo = clase["id_grupo"]
        id_materia = clase["id_materia"]

        # 2️⃣ Actividades
        actividades = await fetch_all("""
            SELECT id_actividad, titulo
            FROM actividad
            WHERE id_clase = %s
            ORDER BY fecha_entrega ASC
        """, (id_clase,))

        # 3️⃣ Alumnos
        alumnos = await fetch_all("""
            SELECT id_estudiante, nombre, apellido, matricula, no_lista
            FROM estudiante
            WHERE id_grupo = %s
            ORDER BY no_lista
        """, (id_grupo,))

        # 4️⃣ Entregas
        entregas_map = {}
        if actividades:
            actividad_ids = [a["id_actividad"] for a in actividades]
            placeholders = ','.join(['%s'] * len(actividad_ids))
            entregas = await fetch_all(f"""
                SELECT id_estudiante, id_actividad, estado, calificacion
                FROM actividad_estudiante
                WHERE id_actividad IN ({placeholders})
            """, tuple(actividad_ids))
            for e in entregas:
                key = f"{e['id_estudiante']}_{e['id_actividad']}"
                entregas_map[key] = e

        # 5️⃣ Crear workbook y hoja ACTIVIDADES
        wb = Workbook()
        ws_act = wb.active
        ws_act.title = nombre_hoja_seguro(f"{materia}-{grupo}", "Act")

        header_act = ["No Lista", "Matrícula", "Nombre", "Grupo"]
        for act in actividades:
            header_act.append(act["titulo"])
            header_act.append(f"Calificación - {act['titulo']}")
        ws_act.append(header_act)
        for cell in ws_act[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rellenar datos y colores de estado
        for alumno in alumnos:
            row_data = [
                alumno["no_lista"],
                alumno["matricula"],
                f"{alumno['apellido']} {alumno['nombre']}",
                grupo
            ]
            for act in actividades:
                key = f"{alumno['id_estudiante']}_{act['id_actividad']}"
                entrega = entregas_map.get(key)
                estado = entrega["estado"] if entrega else "pendiente"
                calificacion = entrega["calificacion"] if entrega and entrega.get("calificacion") is not None else 0
                row_data.append(estado)
                row_data.append(calificacion)
            ws_act.append(row_data)

        # Aplicar colores a celdas de estado
        for row in ws_act.iter_rows(min_row=2, min_col=5, max_col=4 + len(actividades)*2):
            for idx, cell in enumerate(row):
                if idx % 2 == 0:  # solo columnas de estado
                    estado_val = str(cell.value).lower()
                    if estado_val == "entregado":
                        cell.fill = COLOR_ENTREGADO
                    elif estado_val == "pendiente":
                        cell.fill = COLOR_PENDIENTE
                    elif estado_val == "no entregado":
                        cell.fill = COLOR_NO_ENTREGADO
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(bold=True)

        # Ajustar ancho columnas
        for col in ws_act.columns:
            ws_act.column_dimensions[col[0].column_letter].width = 20

        # ===== HOJA ASISTENCIAS =====
        ws_asis = wb.create_sheet(nombre_hoja_seguro(f"{materia}-{grupo}", "Asis"))

        # Obtener asistencias
        asistencias_query = """
            SELECT 
                e.id_estudiante,
                CONCAT(e.nombre, ' ', e.apellido) AS nombre,
                e.matricula,
                a.estado,
                a.fecha
            FROM estudiante e
            JOIN clase c ON c.id_grupo = e.id_grupo
            LEFT JOIN asistencia a ON a.id_estudiante = e.id_estudiante AND a.id_clase = c.id_clase
            WHERE c.id_grupo = %s AND c.id_materia = %s
            ORDER BY e.no_lista, a.fecha
        """
        resultado = await fetch_all(asistencias_query, (id_grupo, id_materia))

        # Fechas únicas de asistencia (solo días que hubo clase)
        fechas_unicas = sorted(list({formato_fecha(r["fecha"]) for r in resultado if r["fecha"]}))

        # Agrupar por estudiante
        estudiantes_map = {}
        for row in resultado:
            id_est = row["id_estudiante"]
            if id_est not in estudiantes_map:
                estudiantes_map[id_est] = {
                    "nombre": row["nombre"],
                    "matricula": row["matricula"],
                    "estados": {}
                }
            if row["fecha"]:
                fecha_str = formato_fecha(row["fecha"])
                estudiantes_map[id_est]["estados"][fecha_str] = row["estado"]

        # Encabezados
        encabezados = ["Nombre", "Matrícula", "Grupo", "Materia"] + fechas_unicas + ["TOTAL P", "TOTAL A", "TOTAL J"]
        ws_asis.append(encabezados)
        for cell in ws_asis[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Llenar filas y conteos
        for est in estudiantes_map.values():
            fila = [
                est["nombre"],
                est["matricula"],
                grupo,
                materia
            ]
            total_P = total_A = total_J = 0
            for f in fechas_unicas:
                estado = est["estados"].get(f)
                letra = (
                    "P" if estado == "presente" else
                    "A" if estado == "ausente" else
                    "J" if estado == "justificante" else "-"
                )
                fila.append(letra)
                if letra == "P": total_P += 1
                elif letra == "A": total_A += 1
                elif letra == "J": total_J += 1
            fila.extend([total_P, total_A, total_J])
            ws_asis.append(fila)

        # Colorear celdas
        color_map = {"P": COLOR_PRESENTE, "A": COLOR_AUSENTE, "J": COLOR_JUSTIFICANTE}
        for row in ws_asis.iter_rows(min_row=2, min_col=5):
            for cell in row:
                letra = str(cell.value).upper()
                if letra in color_map:
                    cell.fill = color_map[letra]
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True)

        # Ajustar ancho columnas
        for col in ws_asis.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws_asis.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

        # Guardar y enviar Excel
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"ClaseCompleto_{grupo}_{materia}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generando el Excel completo: {str(e)}")



@router.get("/excel/profesor/completo/{id_profesor}")
async def reporte_asistencias_profesor(id_profesor: int):
    try:
        # Definir siempre las fechas
        fechaInicio = "2025-08-04"  # inicio de curso
        fechaFin = obtener_fecha_hora_cdmx()["fecha"].strftime("%Y-%m-%d")
        print(f"Generando reporte para profesor {id_profesor}, fechas {fechaInicio} a {fechaFin}")

        # 1️⃣ Obtener clases del profesor
        clases_query = """
            SELECT c.id_clase, c.nombre_clase, m.nombre AS materia, g.nombre AS grupo
            FROM clase c
            LEFT JOIN materia m ON c.id_materia = m.id_materia
            LEFT JOIN grupo g ON c.id_grupo = g.id_grupo
            WHERE c.id_profesor = %s
        """
        clases = await fetch_all(clases_query, (id_profesor,))
        if not clases:
            raise HTTPException(status_code=404, detail="El profesor no tiene clases asignadas.")

        # 2️⃣ Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # quitar hoja por defecto

        for clase in clases:
            nombre_hoja = limpiar_nombre_hoja(f"{clase['materia']} - {clase['grupo']}")
            sheet = wb.create_sheet(nombre_hoja)

            # 2.1 Obtener estudiantes del grupo
            estudiantes_query = """
                SELECT e.id_estudiante, e.nombre, e.apellido, e.matricula, e.no_lista, g.nombre AS grupo
                FROM estudiante e
                JOIN grupo g ON e.id_grupo = g.id_grupo
                WHERE e.id_grupo = (SELECT id_grupo FROM clase WHERE id_clase = %s)
                ORDER BY e.no_lista
            """
            estudiantes = await fetch_all(estudiantes_query, (clase['id_clase'],))

            # 2.2 Obtener asistencias del rango
            asistencias_query = """
                SELECT a.id_estudiante, a.fecha, a.estado
                FROM asistencia a
                WHERE a.id_clase = %s AND a.fecha BETWEEN %s AND %s
            """
            asistencias = await fetch_all(asistencias_query, (clase['id_clase'], fechaInicio, fechaFin))

            # 2.3 Fechas únicas de asistencia
            fechas = sorted(list({formato_fecha(a['fecha']) for a in asistencias}))

            # 2.4 Cabecera
            header = ["No Lista", "Matrícula", "Nombre", "Grupo"] + fechas
            sheet.append(header)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # 2.5 Filas de estudiantes
            for est in estudiantes:
                row = [
                    est['no_lista'],
                    est['matricula'],
                    f"{est['nombre']} {est['apellido']}",
                    est['grupo']
                ]
                for f in fechas:
                    a = next((x for x in asistencias if x['id_estudiante'] == est['id_estudiante'] and formato_fecha(x['fecha']) == f), None)
                    row.append(a['estado'] if a else "—")
                sheet.append(row)

            # 2.6 Colorear celdas
            for r in sheet.iter_rows(min_row=2, min_col=5):
                for cell in r:
                    if cell.value:
                        val = str(cell.value).lower()
                        if val == "presente":
                            cell.fill = COLOR_PRESENTE
                        elif val == "ausente":
                            cell.fill = COLOR_AUSENTE
                        elif val == "justificante":
                            cell.fill = COLOR_JUSTIFICANTE

            # 2.7 Ajustar ancho de columnas
            for col in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                sheet.column_dimensions[col[0].column_letter].width = max(15, max_length)

        # 3️⃣ Enviar Excel
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Reporte_Profesor_{id_profesor}.xlsx"}
        )

    except Exception as e:
        print("❌ ERROR DETALLADO:", e)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generando reporte: {str(e)}")


@router.get("/alumnos/clase/{id_clase}/excel")
async def exportar_excel_alumnos_clase(id_clase: int):
    try:
        # 1️⃣ Obtener info de grupo y materia de la clase
        clase_info_query = """
            SELECT 
                c.id_grupo,
                c.id_materia,
                g.nombre AS nombre_grupo,
                m.nombre AS nombre_materia
            FROM clase c
            JOIN grupo g ON c.id_grupo = g.id_grupo
            JOIN materia m ON c.id_materia = m.id_materia
            WHERE c.id_clase = %s
        """
        clase_info = await fetch_one(clase_info_query, (id_clase,))

        if not clase_info:
            raise HTTPException(status_code=404, detail="Clase no encontrada")

        id_grupo = clase_info["id_grupo"]
        id_materia = clase_info["id_materia"]
        nombre_grupo = clase_info["nombre_grupo"]
        nombre_materia = clase_info["nombre_materia"]

        # 2️⃣ Obtener asistencias del grupo y materia
        asistencias_query = """
            SELECT 
                e.id_estudiante,
                CONCAT(e.nombre, ' ', e.apellido) AS nombre,
                e.matricula,
                a.estado,
                a.fecha
            FROM estudiante e
            JOIN clase c ON c.id_grupo = e.id_grupo
            LEFT JOIN asistencia a ON a.id_estudiante = e.id_estudiante AND a.id_clase = c.id_clase
            WHERE c.id_grupo = %s AND c.id_materia = %s
            ORDER BY e.nombre, e.apellido, a.fecha
        """
        resultado = await fetch_all(asistencias_query, (id_grupo, id_materia))

        # 3️⃣ Extraer fechas únicas
        fechas_unicas = sorted(
            list({formato_fecha(r["fecha"]) for r in resultado if r["fecha"]})
        )

        # 4️⃣ Agrupar por estudiante
        estudiantes_map = {}
        for row in resultado:
            id_est = row["id_estudiante"]
            if id_est not in estudiantes_map:
                estudiantes_map[id_est] = {
                    "nombre": row["nombre"],
                    "matricula": row["matricula"],
                    "estados": {}
                }
            if row["fecha"]:
                fecha_str = formato_fecha(row["fecha"])
                estudiantes_map[id_est]["estados"][fecha_str] = row["estado"]

        # 5️⃣ Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = limpiar_nombre_hoja(f"Asistencias_{nombre_grupo}")

        # 6️⃣ Encabezados
        encabezados = ["Nombre", "Matrícula", "Grupo", "Materia"] + fechas_unicas
        ws.append(encabezados)

        # 7️⃣ Agregar filas con datos
        for est in estudiantes_map.values():
            fila = [
                est["nombre"],
                est["matricula"],
                nombre_grupo,
                nombre_materia,
            ]
            for f in fechas_unicas:
                estado = est["estados"].get(f)
                letra = (
                    "P" if estado == "presente" else
                    "A" if estado == "ausente" else
                    "J" if estado == "justificante" else "A"
                )
                fila.append(letra)
            ws.append(fila)

        # 8️⃣ Estilo de encabezado
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 9️⃣ Colorear celdas según estado
        color_map = {"P": COLOR_PRESENTE, "A": COLOR_AUSENTE, "J": COLOR_JUSTIFICANTE}
        for row in ws.iter_rows(min_row=2, min_col=5):
            for cell in row:
                letra = str(cell.value).upper()
                if letra in color_map:
                    cell.fill = color_map[letra]
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True)

        # 10️⃣ Ajustar ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

        # 11️⃣ Enviar archivo Excel como respuesta
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"Asistencias_{nombre_grupo}_{nombre_materia}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print("❌ ERROR exportando Excel:", e)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error generando el Excel")
