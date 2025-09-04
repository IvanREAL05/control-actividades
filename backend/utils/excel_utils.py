import openpyxl
from io import BytesIO
from typing import List, Dict
from datetime import datetime, time
import logging

# Configurar logger
logger = logging.getLogger(__name__)
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(levelname)s:%(name)s:%(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)


def leer_excel(file: bytes) -> List[Dict]:
    """
    Lee archivo Excel desde bytes y convierte las horas automáticamente
    """
    try:
        workbook = openpyxl.load_workbook(filename=BytesIO(file), data_only=True)
        sheet = workbook.active

        # Obtener headers
        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append(None)

        logger.info(f"Headers encontrados: {headers}")

        data = []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell for cell in row if cell is not None):  # Saltar filas completamente vacías
                continue
                
            clase_dict = {}
            for i in range(len(headers)):
                if i < len(row) and headers[i]:
                    valor = row[i]
                    
                    # Limpiar valores de texto
                    if isinstance(valor, str):
                        valor = valor.strip()
                        if valor == '':
                            valor = None
                    
                    # Convertir horas automáticamente si la columna es de tiempo
                    if headers[i] in ['hora_inicio', 'hora_fin'] and valor is not None:
                        valor = convertir_hora_excel(valor)
                    
                    clase_dict[headers[i]] = valor

            # Solo agregar si tiene datos útiles
            if any(v is not None for v in clase_dict.values()):
                data.append(clase_dict)
                logger.debug(f"Fila {row_num} procesada: {clase_dict}")

        logger.info(f"Excel procesado: {len(data)} filas con datos")
        return data
        
    except Exception as e:
        logger.error(f"Error leyendo archivo Excel: {e}")
        raise


def convertir_hora_excel(hora_valor):
    """
    Convierte diferentes formatos de hora de Excel a datetime.time
    ÚNICA función para convertir horas
    """
    if hora_valor is None:
        return None
        
    try:
        # Si ya es un objeto time
        if isinstance(hora_valor, time):
            return hora_valor
            
        # Si es un datetime (común en Excel)
        if isinstance(hora_valor, datetime):
            return hora_valor.time()
            
        # Si es un string con formato HH:MM:SS o HH:MM
        if isinstance(hora_valor, str):
            hora_limpia = hora_valor.strip()
            if ':' in hora_limpia:
                partes = hora_limpia.split(':')
                if len(partes) >= 2:
                    horas = int(partes[0])
                    minutos = int(partes[1])
                    segundos = int(partes[2]) if len(partes) > 2 else 0
                    return time(horas, minutos, segundos)
            
        # Si es un número (tiempo serial de Excel) - entre 0 y 1
        if isinstance(hora_valor, (int, float)):
            if 0 <= hora_valor <= 1:
                # Es una fracción de día
                total_segundos = int(hora_valor * 24 * 60 * 60)
                horas = total_segundos // 3600
                minutos = (total_segundos % 3600) // 60
                segundos = total_segundos % 60
                return time(horas, minutos, segundos)
            else:
                # Podría ser un timestamp, intentar convertir
                dt = datetime.fromtimestamp(hora_valor)
                return dt.time()
            
        logger.warning(f"Formato de hora no reconocido: {hora_valor} (tipo: {type(hora_valor)})")
        return None
        
    except Exception as e:
        logger.error(f"Error convirtiendo hora {hora_valor}: {e}")
        return None